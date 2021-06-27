import openpyxl


class MasterSheet:
    def __init__(self, work_book):
        self.work_book = work_book
        self.groups = {}
        self.cases = {}
        self.max_cases_in_group = 0
        self.load_group_names()
        self.load_cases()

    def load_group_names(self):
        work_sheet = self.work_book['C']
        for row in range(2, work_sheet.max_row+1):
            self.groups[work_sheet.cell(row=row, column=1).value] = work_sheet.cell(row=row, column=2).value.strip()

    def load_cases(self):
        work_sheet = self.work_book['Master']

        for group in self.groups:
            self.cases[self.groups[group]] = []

        for row in range(2, work_sheet.max_row+1):
            case = {}
            for col in range(1, work_sheet.max_column + 1):
                case[work_sheet.cell(row=1, column=col).value] = work_sheet.cell(row=row, column=col).value
            self.cases[self.groups[case['G']]].append(case)

    def add_and_populate_worksheet_for_variable(self,var_name):
        work_sheet = self.work_book.create_sheet(title=var_name.capitalize())
        len_of_cases_by_group = [len(self.cases[self.groups[group]]) for group in self.groups]
        # populate N column
        work_sheet.cell(row=1, column=1).value = 'N'
        for row in range(2, max(len_of_cases_by_group) + 2):
            work_sheet.cell(row=row, column=1).value = row - 1
        # populate groups columns with variable values
        for col in range(2, len(self.groups) + 2):
            work_sheet.cell(row=1, column=col).value = self.groups[col-1]
            for index, case in enumerate(self.cases[self.groups[col-1]]):
                work_sheet.cell(row=index+2, column=col).value = case[var_name]


def main():
    work_book = openpyxl.load_workbook('sample_test.xlsx')
    x = MasterSheet(work_book)

    variables = ['age', 'weight in kg', 'height in cm', 'surgery duration in min', 'TFR', 'morphine consump', 'fentanyl consump']
    for var in variables:
        x.add_and_populate_worksheet_for_variable(var)

    work_book.save('results.xlsx')

main()
